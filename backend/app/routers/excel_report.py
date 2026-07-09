import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
import json
import base64
import io

def generate_excel_report(output_path, request, df, original_df=None):
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def format_stat(val, digits=3):
        if val is None or val == '': return ''
        try:
            f_val = float(val)
            s = f"{abs(f_val):.{digits}f}"
            if f_val < 0:
                s = f"-{s}"
            if abs(f_val) < 1.0:
                s = s.replace("0.", ".", 1) if not s.startswith("-") else s.replace("-0.", "-.", 1)
            return s
        except (ValueError, TypeError):
            return val

    def format_cell(ws, row, col, value, is_header=False, align='center'):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = thin_border
        if is_header:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        else:
            cell.alignment = center_align if align == 'center' else left_align
        return cell

    # 1. Imported Data (Original)
    if original_df is not None:
        ws_orig = wb.create_sheet("Imported Data")
        columns = list(original_df.columns)
        for c, h in enumerate(columns, 1):
            format_cell(ws_orig, 1, c, h, is_header=True)
        for r_idx, row in enumerate(original_df.itertuples(index=False), start=2):
            for c_idx, val in enumerate(row, start=1):
                ws_orig.cell(row=r_idx, column=c_idx, value=val)

    # 2. Audit Logs (Data Cleansing & Reverse Coding)
    ws_audit = wb.create_sheet("데이터 클린징 이력")
    headers = ["시간", "단계", "액션", "상세 내용"]
    for c, h in enumerate(headers, 1):
        format_cell(ws_audit, 1, c, h, is_header=True)
    ws_audit.column_dimensions['A'].width = 25
    ws_audit.column_dimensions['B'].width = 20
    ws_audit.column_dimensions['C'].width = 30
    ws_audit.column_dimensions['D'].width = 80
    
    if request.auditLogs:
        r_idx = 2
        for log in request.auditLogs:
            if log.get('step') == '인구통계 설정':
                continue # 인구통계는 따로 처리
            details = log.get('details', '')
            action = log.get('action', '')
            details_str = ""
            if isinstance(details, dict):
                if action in ['결측치 수정', '셀 값 수정']:
                    row_id = details.get('rowId', '')
                    col_id = details.get('columnId', '')
                    old_v = details.get('oldValue', '결측치(null)')
                    new_v = details.get('newValue', '결측치(null)')
                    if old_v is None: old_v = '결측치(null)'
                    if new_v is None: new_v = '결측치(null)'
                    details_str = f"[{row_id}번 응답자] '{col_id}' 문항의 값 변경: {old_v} -> {new_v}"
                elif action == '역코딩 적용':
                    cols = details.get('columns', [])
                    details_str = f"문항 {', '.join(cols)} 역코딩 적용 (범위: {details.get('min', 1)}~{details.get('max', 5)})"
                elif action == '항목 삭제':
                    details_str = f"문항 '{details.get('itemId', '')}' 삭제 처리"
                else:
                    details_str = json.dumps(details, ensure_ascii=False)
            else:
                details_str = str(details)
                
            format_cell(ws_audit, r_idx, 1, log.get('timestamp', ''))
            format_cell(ws_audit, r_idx, 2, log.get('step', ''))
            format_cell(ws_audit, r_idx, 3, action)
            format_cell(ws_audit, r_idx, 4, details_str, align='left')
            r_idx += 1
        if r_idx == 2:
            format_cell(ws_audit, 2, 1, "클린징 기록이 없습니다.")
    else:
        format_cell(ws_audit, 2, 1, "클린징 기록이 없습니다.")

    # 3. Demographics Logs
    ws_demo = wb.create_sheet("인구통계 이력")
    for c, h in enumerate(headers, 1):
        format_cell(ws_demo, 1, c, h, is_header=True)
    ws_demo.column_dimensions['A'].width = 25
    ws_demo.column_dimensions['B'].width = 20
    ws_demo.column_dimensions['C'].width = 30
    ws_demo.column_dimensions['D'].width = 80

    if request.auditLogs:
        r_idx = 2
        for log in request.auditLogs:
            if log.get('step') != '인구통계 설정':
                continue
            details = log.get('details', '')
            action = log.get('action', '')
            details_str = ""
            if isinstance(details, dict):
                details_str = json.dumps(details, ensure_ascii=False)
            else:
                details_str = str(details)
                
            format_cell(ws_demo, r_idx, 1, log.get('timestamp', ''))
            format_cell(ws_demo, r_idx, 2, log.get('step', ''))
            format_cell(ws_demo, r_idx, 3, action)
            format_cell(ws_demo, r_idx, 4, details_str, align='left')
            r_idx += 1
        if r_idx == 2:
            format_cell(ws_demo, 2, 1, "인구통계 기록이 없습니다.")
    else:
        format_cell(ws_demo, 2, 1, "인구통계 기록이 없습니다.")

    # 4. Cleansed Data
    ws_data = wb.create_sheet("정제된 데이터 (Data)")
    columns = list(df.columns)
    for c, h in enumerate(columns, 1):
        format_cell(ws_data, 1, c, h, is_header=True)
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws_data.cell(row=r_idx, column=c_idx, value=val)

    # 5. Variable Mapping
    ws_map = wb.create_sheet("변수 매핑 (Variable Mapping)")
    headers = ["유형", "역할", "매핑된 변수 목록"]
    for c, h in enumerate(headers, 1):
        format_cell(ws_map, 1, c, h, is_header=True)
    ws_map.column_dimensions['A'].width = 15
    ws_map.column_dimensions['B'].width = 25
    ws_map.column_dimensions['C'].width = 60
    
    role_names = {
        'iv': '독립변수 (IV)', 'dv': '종속변수 (DV)', 'med': '매개변수 (Mediator)',
        'mod': '조절변수 (Moderator)', 'gen': '인구통계/통제변수'
    }
    
    r_idx = 2
    if 'mappedVars' in request.project_config:
        for role, vars in request.project_config['mappedVars'].items():
            for v in vars:
                format_cell(ws_map, r_idx, 1, role_names.get(role, role))
                format_cell(ws_map, r_idx, 2, v.get('name', ''))
                format_cell(ws_map, r_idx, 3, ", ".join(v.get('itemIds', [])), align='left')
                r_idx += 1

    # 6. Analysis Results & Model
    if request.cachedResults:
        for key, cache in request.cachedResults.items():
            results = cache.get("results", {})
            settings = cache.get("settings", {})
            interpretation = cache.get("interpretation", "")
            
            safe_title = str(key)[:31]
            ws = wb.create_sheet(safe_title)
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            
            curr_row = 1
            if key == 'model':
                ws.title = '모형설계'
                base64_img = settings.get('image', '')
                if base64_img:
                    if ',' in base64_img:
                        base64_img = base64_img.split(',')[1]
                    try:
                        img_data = base64.b64decode(base64_img)
                        img = OpenpyxlImage(io.BytesIO(img_data))
                        ws.add_image(img, 'A1')
                        curr_row += 30 # 이미지 크기에 따라 여백 조절
                    except Exception as e:
                        ws.cell(row=curr_row, column=1, value="이미지 로드 실패")
                        curr_row += 2
                        
                # Print hypotheses
                hypos = results.get('hypotheses', [])
                if hypos:
                    ws.cell(row=curr_row, column=1, value="[자동 생성된 가설]").font = header_font
                    curr_row += 1
                    for i, h in enumerate(hypos, 1):
                        ws.cell(row=curr_row, column=1, value=f"가설 {i}: {h.get('text', '')}")
                        curr_row += 1
                parsed = True
                continue
                
            ws.cell(row=curr_row, column=1, value="[분석 옵션 및 설정]").font = header_font
            curr_row += 1
            for sk, sv in settings.items():
                ws.cell(row=curr_row, column=1, value=str(sk))
                ws.cell(row=curr_row, column=2, value=str(sv))
                curr_row += 1
            curr_row += 1
            
            ws.cell(row=curr_row, column=1, value="[분석 결과 데이터]").font = header_font
            curr_row += 1
            
            parsed = False
            try:
                if key == 'efa':
                    if 'loadings' in results:
                        loadings = results['loadings']
                        factors = list(list(loadings.values())[0].keys()) if loadings else []
                        headers = ["구성요인", "설문문항"] + factors + ["공통성"]
                        for c, h in enumerate(headers, 1): format_cell(ws, curr_row, c, h, is_header=True)
                        curr_row += 1
                        
                        items = list(loadings.keys())
                        
                        # 그룹화: 요인별로 가장 적재값이 큰 문항들 묶기
                        groups = {f: [] for f in factors}
                        for item in items:
                            item_loadings = [abs(loadings[item].get(f, 0)) for f in factors]
                            max_idx = item_loadings.index(max(item_loadings))
                            groups[factors[max_idx]].append(item)
                            
                        # 설정에 따라 그룹 내에서 크기순 정렬 (선택사항, 여기서는 일단 그대로 출력)
                        if settings.get('sortBySize', True):
                            for f in factors:
                                groups[f].sort(key=lambda x: abs(loadings[x].get(f, 0)), reverse=True)

                        for f_idx, f in enumerate(factors):
                            factor_items = groups[f]
                            if not factor_items: continue
                            
                            start_row = curr_row
                            for i, item in enumerate(factor_items):
                                # 구성요인 (첫 행에만)
                                if i == 0:
                                    factor_name = results.get('factorNames', factors)[f_idx]
                                    cell = format_cell(ws, curr_row, 1, f"요인{f_idx+1}\n{factor_name}")
                                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                                else:
                                    format_cell(ws, curr_row, 1, "")
                                    
                                # 설문문항
                                format_cell(ws, curr_row, 2, item)
                                
                                # 적재값
                                c_idx = 3
                                for f_col in factors:
                                    val = loadings[item].get(f_col, '')
                                    format_cell(ws, curr_row, c_idx, format_stat(val, 3))
                                    c_idx += 1
                                    
                                # 공통성
                                comm = results.get('communalities', {}).get(item, '')
                                format_cell(ws, curr_row, c_idx, format_stat(comm, 3))
                                curr_row += 1
                            
                            # 병합 (rowSpan 효과)
                            if len(factor_items) > 1:
                                ws.merge_cells(start_row=start_row, start_column=1, end_row=curr_row-1, end_column=1)
                        
                        var_exp = results.get('variance_explained', [])
                        if var_exp:
                            format_cell(ws, curr_row, 1, "고유값")
                            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
                            format_cell(ws, curr_row, 2, "") # merged
                            c_idx = 3
                            for i, f in enumerate(factors):
                                val = var_exp[i].get('ss_loadings', 0) if i < len(var_exp) else 0
                                format_cell(ws, curr_row, c_idx, round(val, 3) if isinstance(val, float) else val)
                                c_idx += 1
                            format_cell(ws, curr_row, c_idx, "")
                            curr_row += 1
                            
                            format_cell(ws, curr_row, 1, "분산 설명력(%)")
                            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
                            format_cell(ws, curr_row, 2, "") # merged
                            c_idx = 3
                            for i, f in enumerate(factors):
                                val = var_exp[i].get('variance_pct', 0) if i < len(var_exp) else 0
                                format_cell(ws, curr_row, c_idx, round(val, 2) if isinstance(val, float) else val)
                                c_idx += 1
                            format_cell(ws, curr_row, c_idx, "")
                            curr_row += 1
                            
                            format_cell(ws, curr_row, 1, "누적 설명력(%)")
                            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
                            format_cell(ws, curr_row, 2, "") # merged
                            c_idx = 3
                            for i, f in enumerate(factors):
                                val = var_exp[i].get('cumulative_pct', 0) if i < len(var_exp) else 0
                                format_cell(ws, curr_row, c_idx, round(val, 2) if isinstance(val, float) else val)
                                c_idx += 1
                            format_cell(ws, curr_row, c_idx, "")
                            curr_row += 1
                        curr_row += 1
                        
                        # Options summary list
                        opt_extraction = "주성분분석" if settings.get('extraction') == 'pca' else settings.get('extraction', '주성분분석')
                        opt_rotation = "배리맥스" if settings.get('rotation') == 'varimax' else settings.get('rotation', '배리맥스')
                        opt_loading = settings.get('loading', 0.5)
                        opt_communality = settings.get('communality', 0.4)
                        opt_variance = settings.get('variance', 60)
                        opt_kmo_limit = settings.get('kmo', 0.5)
                        
                        options_text = [
                            f"• 요인추출: {opt_extraction}",
                            f"• 요인회전: {opt_rotation}",
                            f"• 요인적재값: {opt_loading} 이상",
                            f"• 공통성: {opt_communality} 이상",
                            f"• 분산설명력: {opt_variance}% 이상",
                            f"• KMO: {opt_kmo_limit} 이상"
                        ]
                        
                        for txt in options_text:
                            ws.cell(row=curr_row, column=1, value=txt)
                            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=len(headers))
                            curr_row += 1
                        
                        curr_row += 1
                        ws.cell(row=curr_row, column=1, value=f"KMO={format_stat(results.get('kmo',0))}, Bartlett's test결과 χ²={round(results.get('bartlett_chi_square',0),3)} (df={results.get('bartlett_df',0)}, p={format_stat(results.get('bartlett_p_value',1))})")
                        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=len(headers))
                        curr_row += 2
                        parsed = True
                        
                elif key == 'reliability':
                    if isinstance(results, list):
                        headers = ["요인", "문항", "항목 삭제시 크론바흐 알파", "크론바흐 알파"]
                        for c, h in enumerate(headers, 1): format_cell(ws, curr_row, c, h, is_header=True)
                        curr_row += 1
                        
                        for data in results:
                            factor = data.get('name', '')
                            ca = data.get('alpha', '')
                            item_stats = data.get('item_stats', [])
                            if item_stats:
                                format_cell(ws, curr_row, 1, factor)
                                format_cell(ws, curr_row, 2, item_stats[0].get('item', ''))
                                val = item_stats[0].get('alpha_if_deleted', '')
                                format_cell(ws, curr_row, 3, round(val, 3) if isinstance(val, float) else val)
                                format_cell(ws, curr_row, 4, round(ca, 3) if isinstance(ca, float) else ca)
                                curr_row += 1
                                for stat in item_stats[1:]:
                                    format_cell(ws, curr_row, 1, factor)
                                    format_cell(ws, curr_row, 2, stat.get('item', ''))
                                    val = stat.get('alpha_if_deleted', '')
                                    format_cell(ws, curr_row, 3, round(val, 3) if isinstance(val, float) else val)
                                    format_cell(ws, curr_row, 4, "")
                                    curr_row += 1
                        parsed = True

                elif key == 'frequency':
                    for col_name, freqs in results.items():
                        ws.cell(row=curr_row, column=1, value=f"[{col_name}]").font = header_font
                        curr_row += 1
                        headers = ["범주 (Category)", "빈도 (N)", "퍼센트 (%)"]
                        for c, h in enumerate(headers, 1): format_cell(ws, curr_row, c, h, is_header=True)
                        curr_row += 1
                        for f in freqs:
                            format_cell(ws, curr_row, 1, f.get('category', ''))
                            format_cell(ws, curr_row, 2, f.get('count', 0))
                            format_cell(ws, curr_row, 3, round(f.get('percent', 0), 1))
                            curr_row += 1
                        curr_row += 1
                    parsed = True

                elif key == 'ttest':
                    if 'results' in results:
                        for dv_name, tests in results['results'].items():
                            ws.cell(row=curr_row, column=1, value=f"종속변수: {dv_name}").font = header_font
                            curr_row += 1
                            headers = ["독립변수", "집단", "N", "평균", "표준편차", "t/F", "p-value", "사후검정"]
                            for c, h in enumerate(headers, 1): format_cell(ws, curr_row, c, h, is_header=True)
                            curr_row += 1
                            for test in tests:
                                iv_name = test.get('iv_name', '')
                                method = test.get('method', '')
                                groups = test.get('groups', [])
                                has_posthoc = 'posthoc' in test and test['posthoc']
                                
                                if groups:
                                    for i, g in enumerate(groups):
                                        format_cell(ws, curr_row, 1, iv_name if i == 0 else "")
                                        format_cell(ws, curr_row, 2, g.get('group_name'))
                                        format_cell(ws, curr_row, 3, g.get('n'))
                                        format_cell(ws, curr_row, 4, round(g.get('mean'), 3) if isinstance(g.get('mean'), float) else g.get('mean'))
                                        format_cell(ws, curr_row, 5, round(g.get('sd'), 3) if isinstance(g.get('sd'), float) else g.get('sd'))
                                        
                                        if i == 0:
                                            stat_val = test.get('t_stat') if method == 't-test' else test.get('f_stat')
                                            format_cell(ws, curr_row, 6, round(stat_val, 3) if isinstance(stat_val, float) else stat_val)
                                            p_val = test.get('p_value')
                                            format_cell(ws, curr_row, 7, round(p_val, 3) if isinstance(p_val, float) else p_val)
                                            if has_posthoc:
                                                format_cell(ws, curr_row, 8, ", ".join(test['posthoc']) if isinstance(test['posthoc'], list) else str(test['posthoc']))
                                            else:
                                                format_cell(ws, curr_row, 8, "")
                                        else:
                                            format_cell(ws, curr_row, 6, "")
                                            format_cell(ws, curr_row, 7, "")
                                            format_cell(ws, curr_row, 8, "")
                                        curr_row += 1
                            curr_row += 1
                        parsed = True

                elif key == 'correlation':
                    if 'factor_names' in results:
                        names = results['factor_names']
                        mat_r = results.get('matrix_r', [])
                        mat_p = results.get('matrix_p', [])
                        
                        headers = ["구분"] + names
                        for c, h in enumerate(headers, 1): format_cell(ws, curr_row, c, h, is_header=True)
                        curr_row += 1
                        
                        for i, name in enumerate(names):
                            format_cell(ws, curr_row, 1, name)
                            for j in range(len(names)):
                                if i == j:
                                    format_cell(ws, curr_row, j+2, "1")
                                elif j < i:
                                    r = mat_r[i][j]
                                    p = mat_p[i][j]
                                    if r is None:
                                        format_cell(ws, curr_row, j+2, "-")
                                    else:
                                        stars = ""
                                        if p is not None:
                                            if p < 0.001: stars = "***"
                                            elif p < 0.01: stars = "**"
                                            elif p < 0.05: stars = "*"
                                        val_str = f"{r:.3f}".replace("0.", ".") + stars
                                        format_cell(ws, curr_row, j+2, val_str)
                                else:
                                    format_cell(ws, curr_row, j+2, "")
                            curr_row += 1
                        parsed = True

                elif key.startswith('regression'):
                    if 'models' in results:
                        for idx, model in enumerate(results['models'], 1):
                            ws.cell(row=curr_row, column=1, value=f"[Model {idx}] 종속변수: {model.get('dv_name', '')}").font = header_font
                            curr_row += 1
                            
                            headers = ["R", "R²", "Adj R²", "F값", "유의확률(p)"]
                            for c, h in enumerate(headers, 1): format_cell(ws, curr_row, c, h, is_header=True)
                            curr_row += 1
                            
                            r_squared = model.get('r_squared', 0)
                            r_val = r_squared ** 0.5 if r_squared > 0 else 0
                            
                            format_cell(ws, curr_row, 1, f"{r_val:.3f}".replace("0.", "."))
                            format_cell(ws, curr_row, 2, f"{r_squared:.3f}".replace("0.", "."))
                            format_cell(ws, curr_row, 3, f"{model.get('adj_r_squared', 0):.3f}".replace("0.", "."))
                            format_cell(ws, curr_row, 4, round(model.get('f_value', 0), 3))
                            format_cell(ws, curr_row, 5, round(model.get('f_p_value', 1), 3))
                            curr_row += 2
                            
                            headers = ["독립변수", "B", "표준오차", "Beta", "t값", "유의확률(p)", "VIF"]
                            for c, h in enumerate(headers, 1): format_cell(ws, curr_row, c, h, is_header=True)
                            curr_row += 1
                            
                            for coef in model.get('coefficients', []):
                                format_cell(ws, curr_row, 1, coef.get('name', ''))
                                format_cell(ws, curr_row, 2, round(coef.get('B', 0), 3))
                                format_cell(ws, curr_row, 3, round(coef.get('SE', 0), 3))
                                format_cell(ws, curr_row, 4, round(coef.get('beta', 0), 3) if coef.get('beta') is not None else "")
                                format_cell(ws, curr_row, 5, round(coef.get('t', 0), 3))
                                format_cell(ws, curr_row, 6, round(coef.get('p', 1), 3))
                                format_cell(ws, curr_row, 7, round(coef.get('vif', 0), 3) if coef.get('vif') is not None else "")
                                curr_row += 1
                            curr_row += 1
                        parsed = True
                        
                elif key.startswith('mediation') or key.startswith('moderation') or key.startswith('mod_med'):
                    res_list = results.get('results', [])
                    if res_list:
                        for dvRes in res_list:
                            dv_name = dvRes.get('dv_name', '')
                            ws.cell(row=curr_row, column=1, value=f"종속변수: {dv_name}").font = header_font
                            curr_row += 1
                            
                            models = dvRes.get('models', [])
                            if models:
                                if 'model1' in models[0]:
                                    for m in models:
                                        iv_name = m.get('iv_name', 'IV')
                                        mod_name = m.get('mod_name', 'MOD')
                                        ws.cell(row=curr_row, column=1, value=f"{iv_name} * {mod_name} -> {dv_name}").font = header_font
                                        curr_row += 1
                                        headers = ["구분", "모델1 (B)", "모델2 (B)", "모델3 (B)"]
                                        for c, h in enumerate(headers, 1): format_cell(ws, curr_row, c, h, is_header=True)
                                        curr_row += 1
                                        
                                        rows_dict = {}
                                        for c_idx, mod_key in enumerate(['model1', 'model2', 'model3'], 2):
                                            if mod_key in m:
                                                for coef in m[mod_key].get('coefficients', []):
                                                    name = coef.get('name')
                                                    if name not in rows_dict: rows_dict[name] = ["", "", ""]
                                                    p = coef.get('p_value', 1)
                                                    stars = "***" if p<0.001 else "**" if p<0.01 else "*" if p<0.05 else ""
                                                    rows_dict[name][c_idx-2] = f"{coef.get('B',0):.3f}{stars}"
                                        
                                        for name, vals in rows_dict.items():
                                            format_cell(ws, curr_row, 1, name)
                                            for i, v in enumerate(vals): format_cell(ws, curr_row, i+2, v)
                                            curr_row += 1
                                        
                                        format_cell(ws, curr_row, 1, "R²")
                                        for i, mod_key in enumerate(['model1', 'model2', 'model3']):
                                            val = m.get(mod_key, {}).get('model_summary', {}).get('r_square', '')
                                            val_str = f"{val:.3f}".replace("0.", ".") if isinstance(val, (int, float)) else val
                                            format_cell(ws, curr_row, i+2, val_str)
                                        curr_row += 1
                                        
                                        format_cell(ws, curr_row, 1, "F값")
                                        for i, mod_key in enumerate(['model1', 'model2', 'model3']):
                                            val = m.get(mod_key, {}).get('model_summary', {}).get('f_stat', '')
                                            format_cell(ws, curr_row, i+2, round(val, 3) if isinstance(val, float) else val)
                                        curr_row += 2
                                elif 'step1' in models[0] or 'm_model' in models[0]:
                                    for m in models:
                                        is_modmed = 'm_model' in m
                                        m_key = 'm_model' if is_modmed else 'step2'
                                        y_key = 'y_model' if is_modmed else 'step3'
                                        med_name = m.get('med_name', 'M')
                                        ws.cell(row=curr_row, column=1, value=f"매개변수: {med_name}").font = header_font
                                        curr_row += 1
                                        
                                        headers = ["구분", "매개변수모형(M)", "종속변수모형(Y)"]
                                        for c, h in enumerate(headers, 1): format_cell(ws, curr_row, c, h, is_header=True)
                                        curr_row += 1
                                        
                                        rows_dict = {}
                                        for c_idx, mod_key in enumerate([m_key, y_key], 2):
                                            if mod_key in m:
                                                for coef in m[mod_key].get('coefficients', []):
                                                    name = coef.get('name')
                                                    if name not in rows_dict: rows_dict[name] = ["", ""]
                                                    p = coef.get('p_value', 1)
                                                    stars = "***" if p<0.001 else "**" if p<0.01 else "*" if p<0.05 else ""
                                                    rows_dict[name][c_idx-2] = f"{coef.get('B',0):.3f}{stars}"
                                                    
                                        for name, vals in rows_dict.items():
                                            format_cell(ws, curr_row, 1, name)
                                            for i, v in enumerate(vals): format_cell(ws, curr_row, i+2, v)
                                            curr_row += 1
                                            
                                        format_cell(ws, curr_row, 1, "R²")
                                        for i, mod_key in enumerate([m_key, y_key]):
                                            val = m.get(mod_key, {}).get('model_summary', {}).get('r_square', '')
                                            format_cell(ws, curr_row, i+2, round(val, 3) if isinstance(val, float) else val)
                                        curr_row += 1
                                        
                                        format_cell(ws, curr_row, 1, "F값")
                                        for i, mod_key in enumerate([m_key, y_key]):
                                            val = m.get(mod_key, {}).get('model_summary', {}).get('f_stat', '')
                                            format_cell(ws, curr_row, i+2, round(val, 3) if isinstance(val, float) else val)
                                        curr_row += 2
                                        
                                        if 'sobel' in m:
                                            sobel = m['sobel']
                                            if sobel:
                                                ws.cell(row=curr_row, column=1, value="[Sobel Test]").font = header_font
                                                curr_row += 1
                                                format_cell(ws, curr_row, 1, "Z", is_header=True)
                                                format_cell(ws, curr_row, 2, "p-value", is_header=True)
                                                curr_row += 1
                                                format_cell(ws, curr_row, 1, round(sobel.get('z',0),3))
                                                format_cell(ws, curr_row, 2, round(sobel.get('p',1),3))
                                                curr_row += 2
                                        if 'bootstrap' in m:
                                            boot = m['bootstrap']
                                            if boot:
                                                ws.cell(row=curr_row, column=1, value="[Bootstrap]").font = header_font
                                                curr_row += 1
                                                format_cell(ws, curr_row, 1, "Effect", is_header=True)
                                                format_cell(ws, curr_row, 2, "SE", is_header=True)
                                                format_cell(ws, curr_row, 3, "LLCI", is_header=True)
                                                format_cell(ws, curr_row, 4, "ULCI", is_header=True)
                                                curr_row += 1
                                                format_cell(ws, curr_row, 1, round(boot.get('effect',0),3))
                                                format_cell(ws, curr_row, 2, round(boot.get('se',0),3))
                                                format_cell(ws, curr_row, 3, round(boot.get('llci',0),3))
                                                format_cell(ws, curr_row, 4, round(boot.get('ulci',0),3))
                                                curr_row += 2
                                        
                                        if 'conditional_effects' in m and m['conditional_effects']:
                                            ws.cell(row=curr_row, column=1, value="[Conditional Effects]").font = header_font
                                            curr_row += 1
                                            format_cell(ws, curr_row, 1, "조절변수 값", is_header=True)
                                            format_cell(ws, curr_row, 2, "Effect", is_header=True)
                                            format_cell(ws, curr_row, 3, "SE", is_header=True)
                                            format_cell(ws, curr_row, 4, "t", is_header=True)
                                            format_cell(ws, curr_row, 5, "p", is_header=True)
                                            curr_row += 1
                                            for ce in m['conditional_effects']:
                                                format_cell(ws, curr_row, 1, round(ce.get('mod_value',0),3))
                                                format_cell(ws, curr_row, 2, round(ce.get('effect',0),3))
                                                format_cell(ws, curr_row, 3, round(ce.get('se',0),3))
                                                format_cell(ws, curr_row, 4, round(ce.get('t',0),3))
                                                format_cell(ws, curr_row, 5, round(ce.get('p',1),3))
                                                curr_row += 1
                                            curr_row += 1
                                            
                                        if 'index_of_moderated_mediation' in m and m['index_of_moderated_mediation']:
                                            ws.cell(row=curr_row, column=1, value="[Index of Moderated Mediation]").font = header_font
                                            curr_row += 1
                                            idx_m = m['index_of_moderated_mediation']
                                            format_cell(ws, curr_row, 1, "Index", is_header=True)
                                            format_cell(ws, curr_row, 2, "SE", is_header=True)
                                            format_cell(ws, curr_row, 3, "LLCI", is_header=True)
                                            format_cell(ws, curr_row, 4, "ULCI", is_header=True)
                                            curr_row += 1
                                            format_cell(ws, curr_row, 1, round(idx_m.get('index',0),3))
                                            format_cell(ws, curr_row, 2, round(idx_m.get('se',0),3))
                                            format_cell(ws, curr_row, 3, round(idx_m.get('llci',0),3))
                                            format_cell(ws, curr_row, 4, round(idx_m.get('ulci',0),3))
                                            curr_row += 2
                        parsed = True
            except Exception as e:
                import traceback
                traceback.print_exc()
                parsed = False
                
            if interpretation:
                ws.cell(row=curr_row, column=1, value="[분석 결과 해석 (APA 가이드라인)]").font = header_font
                curr_row += 1
                lines = interpretation.split('\n')
                for line in lines:
                    ws.cell(row=curr_row, column=1, value=line)
                    curr_row += 1
                curr_row += 1

            if not parsed:
                # 에러 또는 기타 알 수 없는 포맷인 경우 원시 JSON 덤프
                ws.cell(row=curr_row, column=1, value=json.dumps(results, ensure_ascii=False, indent=2))
                
    wb.save(output_path)
